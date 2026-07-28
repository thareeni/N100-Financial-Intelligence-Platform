"""
Stock Screener Engine & Composite Quality Scoring Module.
Loads screener_config.yaml, applies multi-criteria filters on latest financial_ratios + market_cap data,
computes P10/P90 Winsorised composite quality scores (0-100), and exports screener_output.xlsx.
"""

import os
import sqlite3
import yaml
import logging
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("ScreenerEngine")

DB_PATH = os.getenv("DB_PATH", "nifty100.db")
CONFIG_PATH = os.path.join("config", "screener_config.yaml")


class ScreenerEngine:
    def __init__(self, db_path: str = DB_PATH, config_path: str = CONFIG_PATH):
        self.db_path = db_path
        self.config_path = config_path
        self.config = self.load_config()
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)

    def load_config(self) -> Dict:
        if os.path.exists(self.config_path):
            with open(self.config_path, "r", encoding="utf-8") as f:
                return yaml.safe_load(f)
        logger.warning(f"Config file {self.config_path} not found. Using default empty config.")
        return {"preset_screeners": {}, "composite_score_weights": {}}

    def get_latest_universe_df(self) -> pd.DataFrame:
        """
        Queries nifty100.db for the latest valid annual financial year profile per company
        across financial_ratios, market_cap, sectors, companies, and profitandloss.
        """
        conn = sqlite3.connect(self.db_path)
        
        # Query latest valid annual financial_ratios per company
        fr_sql = """
            SELECT fr.*, c.company_name, s.broad_sector, s.sub_sector, pl.sales, pl.net_profit
            FROM financial_ratios fr
            JOIN (
                SELECT company_id, MAX(year) as max_yr
                FROM financial_ratios
                WHERE year != 'PARSE_ERROR' AND return_on_equity_pct IS NOT NULL
                GROUP BY company_id
            ) latest ON fr.company_id = latest.company_id AND fr.year = latest.max_yr
            JOIN companies c ON fr.company_id = c.id
            LEFT JOIN sectors s ON fr.company_id = s.company_id
            LEFT JOIN profitandloss pl ON fr.company_id = pl.company_id AND fr.year = pl.year
        """
        fr_df = pd.read_sql_query(fr_sql, conn)

        # Query latest market_cap per company
        mc_sql = """
            SELECT mc.company_id, mc.pe_ratio, mc.pb_ratio, mc.ev_ebitda, mc.dividend_yield_pct, mc.market_cap_crore
            FROM market_cap mc
            JOIN (
                SELECT company_id, MAX(year) as max_yr
                FROM market_cap
                GROUP BY company_id
            ) latest ON mc.company_id = latest.company_id AND mc.year = latest.max_yr
        """
        mc_df = pd.read_sql_query(mc_sql, conn)

        conn.close()

        # Merge ratio data with market cap valuation multiples
        df = pd.merge(fr_df, mc_df, on="company_id", how="left")
        
        # Fill missing valuation multiples with median defaults if missing
        if "pe_ratio" in df.columns:
            df["pe_ratio"] = df["pe_ratio"].fillna(df["pe_ratio"].median())
        if "pb_ratio" in df.columns:
            df["pb_ratio"] = df["pb_ratio"].fillna(df["pb_ratio"].median())
        if "dividend_yield_pct" in df.columns:
            df["dividend_yield_pct"] = df["dividend_yield_pct"].fillna(df["dividend_yield_pct"].median())

        # Calculate composite quality score for universe
        df = self.calculate_composite_score(df)
        return df

    def calculate_composite_score(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Calculates 0-100 Composite Quality Score using P10/P90 Winsorisation:
        - Profitability (35%): ROE (15%), ROCE (10%), NPM (10%)
        - Cash Quality (30%): FCF (15%), CFO/PAT (10%), FCF > 0 flag (5%)
        - Growth (20%): Revenue CAGR 5yr (10%), PAT CAGR 5yr (10%)
        - Leverage (15%): D/E score (10%), ICR score (5%)
        """
        res_df = df.copy()

        def winsorise_and_scale(series: pd.Series, invert: bool = False) -> pd.Series:
            clean_s = pd.to_numeric(series, errors="coerce").fillna(0.0)
            if len(clean_s) == 0:
                return pd.Series(dtype=float, index=series.index)
            p10 = np.percentile(clean_s, 10)
            p90 = np.percentile(clean_s, 90)
            
            clipped = np.clip(clean_s, p10, p90)
            if p90 == p10:
                scaled = np.zeros_like(clipped) + 50.0
            else:
                scaled = ((clipped - p10) / (p90 - p10)) * 100.0

            if invert:
                scaled = 100.0 - scaled
            return pd.Series(scaled, index=series.index)

        # 1. Profitability Scores
        roe_score = winsorise_and_scale(res_df["return_on_equity_pct"])
        roce_score = winsorise_and_scale(res_df["return_on_capital_employed_pct"])
        npm_score = winsorise_and_scale(res_df["net_profit_margin_pct"])
        prof_score = 0.15 * roe_score + 0.10 * roce_score + 0.10 * npm_score

        # 2. Cash Quality Scores
        fcf_score = winsorise_and_scale(res_df["free_cash_flow_cr"])
        cfo_pat_score = winsorise_and_scale(res_df["cfo_pat_ratio"])
        fcf_flag_score = (res_df["free_cash_flow_cr"] > 0).astype(float) * 100.0
        cash_score = 0.15 * fcf_score + 0.10 * cfo_pat_score + 0.05 * fcf_flag_score

        # 3. Growth Scores
        rev_cagr_score = winsorise_and_scale(res_df["revenue_cagr_5yr"].fillna(0))
        pat_cagr_score = winsorise_and_scale(res_df["pat_cagr_5yr"].fillna(0))
        growth_score = 0.10 * rev_cagr_score + 0.10 * pat_cagr_score

        # 4. Leverage Scores
        de_series = pd.to_numeric(res_df["debt_to_equity"], errors="coerce").fillna(0.0)
        de_score = winsorise_and_scale(de_series, invert=True)
        fin_mask = res_df["broad_sector"] == "Financials"
        de_score[fin_mask] = 80.0  # Financial sector carve-out

        icr_series = pd.to_numeric(res_df["interest_coverage"], errors="coerce").fillna(100.0)
        icr_score = winsorise_and_scale(icr_series)
        lev_score = 0.10 * de_score + 0.05 * icr_score

        composite = prof_score + cash_score + growth_score + lev_score
        res_df["composite_quality_score"] = composite.round(2)
        return res_df

    def apply_filters(self, df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
        """
        Applies up to 15 criteria filters on company DataFrame.
        Special rules:
        - Financial sector companies skip max_de filter.
        - Debt-free companies treat ICR as infinity/safe (passing min_icr).
        """
        filtered = df.copy()

        for f_key, f_val in filters.items():
            if f_val is None:
                continue

            if f_key == "min_roe":
                filtered = filtered[filtered["return_on_equity_pct"] >= f_val]
            elif f_key == "max_de":
                is_fin = filtered["broad_sector"] == "Financials"
                passes_de = filtered["debt_to_equity"] <= f_val
                filtered = filtered[is_fin | passes_de]
            elif f_key == "min_fcf":
                filtered = filtered[filtered["free_cash_flow_cr"] >= f_val]
            elif f_key == "min_revenue_cagr_5yr":
                filtered = filtered[filtered["revenue_cagr_5yr"].fillna(0) >= f_val]
            elif f_key == "min_revenue_cagr_3yr":
                filtered = filtered[filtered["revenue_cagr_3yr"].fillna(0) >= f_val]
            elif f_key == "min_pat_cagr_5yr":
                filtered = filtered[filtered["pat_cagr_5yr"].fillna(0) >= f_val]
            elif f_key == "min_opm":
                filtered = filtered[filtered["operating_profit_margin_pct"].fillna(0) >= f_val]
            elif f_key == "max_pe":
                filtered = filtered[(filtered["pe_ratio"] > 0) & (filtered["pe_ratio"] <= f_val)]
            elif f_key == "max_pb":
                filtered = filtered[(filtered["pb_ratio"] > 0) & (filtered["pb_ratio"] <= f_val)]
            elif f_key == "min_dividend_yield":
                filtered = filtered[filtered["dividend_yield_pct"] >= f_val]
            elif f_key == "min_icr":
                # Debt-free companies (interest_coverage is None/NaN or debt_to_equity == 0) treat ICR as infinity/safe
                is_debt_free = filtered["interest_coverage"].isna() | (filtered["debt_to_equity"] == 0)
                passes_icr = filtered["interest_coverage"] >= f_val
                filtered = filtered[is_debt_free | passes_icr]
            elif f_key == "min_market_cap":
                filtered = filtered[filtered["market_cap_crore"].fillna(0) >= f_val]
            elif f_key == "min_net_profit":
                filtered = filtered[filtered["net_profit"].fillna(0) >= f_val]
            elif f_key == "min_eps_cagr_5yr":
                filtered = filtered[filtered["eps_cagr_5yr"].fillna(0) >= f_val]
            elif f_key == "min_asset_turnover":
                filtered = filtered[filtered["asset_turnover"].fillna(0) >= f_val]
            elif f_key == "max_dividend_payout":
                filtered = filtered[filtered["dividend_payout_ratio_pct"].fillna(0) <= f_val]
            elif f_key == "min_sales":
                filtered = filtered[filtered["sales"].fillna(0) >= f_val]

        return filtered.sort_values(by="composite_quality_score", ascending=False).reset_index(drop=True)

    def run_preset_screener(self, preset_name: str, universe_df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
        if universe_df is None:
            universe_df = self.get_latest_universe_df()

        presets = self.config.get("preset_screeners", {})
        if preset_name not in presets:
            logger.error(f"Preset screener '{preset_name}' not defined in config.")
            return pd.DataFrame()

        preset_info = presets[preset_name]
        filters = preset_info.get("filters", {})
        res = self.apply_filters(universe_df, filters)
        logger.info(f"Preset Screener '{preset_name}' returned {len(res)} matching companies.")
        return res

    def export_screener_excel(self, output_filename: str = "screener_output.xlsx") -> str:
        universe_df = self.get_latest_universe_df()
        presets = self.config.get("preset_screeners", {})

        output_path = os.path.join(self.output_dir, output_filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        target_kpi_cols = [
            "company_id",
            "company_name",
            "broad_sector",
            "composite_quality_score",
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "net_profit_margin_pct",
            "operating_profit_margin_pct",
            "debt_to_equity",
            "interest_coverage",
            "free_cash_flow_cr",
            "cash_from_operations_cr",
            "cfo_pat_ratio",
            "capex_intensity_pct",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "pe_ratio",
            "pb_ratio",
            "ev_ebitda",
            "dividend_yield_pct"
        ]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        for p_name in presets.keys():
            res_df = self.run_preset_screener(p_name, universe_df)
            
            avail_cols = [c for c in target_kpi_cols if c in res_df.columns]
            export_df = res_df[avail_cols].copy()

            ws = wb.create_sheet(title=p_name[:31])
            
            ws.append(avail_cols)
            for col_num in range(1, len(avail_cols) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r_idx, row in export_df.iterrows():
                row_vals = [row[c] for c in avail_cols]
                ws.append(row_vals)
                current_row = ws.max_row
                
                score_val = row.get("composite_quality_score", 0)
                if score_val and score_val >= 70:
                    for col_num in range(1, len(avail_cols) + 1):
                        ws.cell(row=current_row, column=col_num).fill = green_fill
                elif score_val and score_val < 40:
                    for col_num in range(1, len(avail_cols) + 1):
                        ws.cell(row=current_row, column=col_num).fill = red_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        logger.info(f"Exported Screener Excel Report to {output_path}.")
        return output_path


if __name__ == "__main__":
    engine = ScreenerEngine()
    engine.export_screener_excel()
