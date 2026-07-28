"""
Investment Score Engine & Investment Intelligence Report Exporter.
Synthesizes Quality (30%), Growth (25%), Value (20%), Health (15%), Momentum (10%)
into a 0-100 Investment Score and rating (Strong Buy, Buy, Hold, Avoid).
Populates investment_scores SQLite table and exports output/investment_intelligence.xlsx.
"""

import os
import sqlite3
import logging
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from dotenv import load_dotenv

from src.screener.engine import ScreenerEngine
from src.analytics.financial_health import FinancialHealthEngine
from src.analytics.dupont import DuPontEngine
from src.analytics.valuation import ValuationEngine

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("InvestmentScoreEngine")

DB_PATH = os.getenv("DB_PATH", "nifty100.db")


def calculate_investment_rating(score: float) -> str:
    """Assigns investment rating based on 0-100 investment score."""
    if score >= 75.0:
        return "Strong Buy"
    elif score >= 60.0:
        return "Buy"
    elif score >= 45.0:
        return "Hold"
    else:
        return "Avoid"


class InvestmentScoreEngine:
    def __init__(self, db_path: str = DB_PATH):
        self.db_path = db_path
        self.output_dir = "output"
        os.makedirs(self.output_dir, exist_ok=True)

    def run(self) -> pd.DataFrame:
        """
        Runs Health, DuPont, and Valuation engines, then combines all dimensions
        into investment_scores table and generates investment_intelligence.xlsx.
        """
        fh_engine = FinancialHealthEngine(self.db_path)
        fh_df = fh_engine.run()

        dp_engine = DuPontEngine(self.db_path)
        dp_df = dp_engine.run()

        val_engine = ValuationEngine(self.db_path)
        val_df = val_engine.run()

        scr_engine = ScreenerEngine(self.db_path)
        scr_df = scr_engine.get_latest_universe_df()

        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA foreign_keys = ON;")

        # Ensure investment_scores table exists
        conn.execute("""
            CREATE TABLE IF NOT EXISTS investment_scores (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id VARCHAR NOT NULL,
                year VARCHAR NOT NULL,
                quality_score NUMERIC NOT NULL,
                growth_score NUMERIC NOT NULL,
                value_score NUMERIC NOT NULL,
                health_score NUMERIC NOT NULL,
                momentum_score NUMERIC NOT NULL,
                investment_score NUMERIC NOT NULL,
                investment_rating VARCHAR NOT NULL,
                UNIQUE (company_id, year),
                FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE
            );
        """)
        conn.commit()

        # Get latest stock price returns for momentum
        sp_sql = """
            SELECT company_id,
                   ((MAX(close_price) - MIN(close_price)) / MIN(close_price) * 100.0) as price_return_1yr
            FROM stock_prices
            GROUP BY company_id
        """
        sp_df = pd.read_sql_query(sp_sql, conn)
        conn.close()

        merged = pd.merge(scr_df, val_df, on=["company_id", "year"], how="left")
        merged = pd.merge(merged, fh_df, on=["company_id", "year"], how="left")
        merged = pd.merge(merged, sp_df, on="company_id", how="left")

        records = []
        for _, row in merged.iterrows():
            cid = row["company_id"]
            yr = row["year"]

            # 1. Quality Score (30%)
            q_val = row.get("composite_quality_score")
            q_score = float(q_val) if pd.notna(q_val) else 50.0

            # 2. Growth Score (25%)
            rev_cagr = row.get("revenue_cagr_5yr") if pd.notna(row.get("revenue_cagr_5yr")) else 10.0
            pat_cagr = row.get("pat_cagr_5yr") if pd.notna(row.get("pat_cagr_5yr")) else 10.0
            g_score = min(max((rev_cagr * 2.5 + pat_cagr * 2.5), 0.0), 100.0)

            # 3. Value Score (20%)
            v_val = row.get("valuation_score")
            v_score = float(v_val) if pd.notna(v_val) else 50.0

            # 4. Health Score (15%)
            z_val = row.get("altman_z_score") if pd.notna(row.get("altman_z_score")) else 2.5
            if z_val > 2.99:
                h_score = min(70.0 + (z_val * 5.0), 100.0)
            elif z_val >= 1.81:
                h_score = 50.0 + (z_val * 10.0)
            else:
                h_score = max(z_val * 20.0, 10.0)

            # 5. Momentum Score (10%)
            mom_ret = row.get("price_return_1yr") if pd.notna(row.get("price_return_1yr")) else 15.0
            m_score = min(max(50.0 + (mom_ret * 0.5), 0.0), 100.0)

            inv_score = round(0.30 * q_score + 0.25 * g_score + 0.20 * v_score + 0.15 * h_score + 0.10 * m_score, 2)
            inv_rating = calculate_investment_rating(inv_score)

            records.append({
                "company_id": cid,
                "year": yr,
                "quality_score": round(q_score, 2),
                "growth_score": round(g_score, 2),
                "value_score": round(v_score, 2),
                "health_score": round(h_score, 2),
                "momentum_score": round(m_score, 2),
                "investment_score": inv_score,
                "investment_rating": inv_rating
            })

        res_df = pd.DataFrame(records)
        res_df = res_df.drop_duplicates(subset=["company_id", "year"]).reset_index(drop=True)
        
        conn = sqlite3.connect(self.db_path)
        conn.execute("DELETE FROM investment_scores;")
        res_df.to_sql("investment_scores", conn, if_exists="append", index=False)
        conn.commit()
        conn.close()

        logger.info(f"Populated investment_scores table with {len(res_df)} rows.")

        self.export_investment_intelligence_excel()
        return res_df

    def export_investment_intelligence_excel(self, output_filename: str = "investment_intelligence.xlsx") -> str:
        """
        Exports formatted 5-sheet Excel report:
        1. Investment Ranking (Top 20 opportunities highlighted)
        2. Financial Health (Altman Z & Beneish M)
        3. Valuation Analysis (Valuation metrics)
        4. DuPont Analysis (3-stage ROE)
        5. Risk Flags (High manipulation risk / Distress zone)
        """
        conn = sqlite3.connect(self.db_path)

        # 1. Investment Ranking Query
        inv_sql = """
            SELECT c.id as Ticker, c.company_name as Company, s.broad_sector as Sector,
                   ins.investment_score, ins.investment_rating,
                   ins.quality_score, ins.growth_score, ins.value_score, ins.health_score, ins.momentum_score,
                   fr.return_on_equity_pct as ROE_pct, fr.debt_to_equity as DE_ratio, fr.free_cash_flow_cr as FCF_cr,
                   mc.pe_ratio as PE_ratio, mc.market_cap_crore as MarketCap_cr
            FROM investment_scores ins
            JOIN companies c ON ins.company_id = c.id
            LEFT JOIN sectors s ON ins.company_id = s.company_id
            LEFT JOIN (
                SELECT fr_inner.*
                FROM financial_ratios fr_inner
                JOIN (
                    SELECT company_id, MAX(year) as max_yr FROM financial_ratios WHERE year != 'PARSE_ERROR' GROUP BY company_id
                ) latest ON fr_inner.company_id = latest.company_id AND fr_inner.year = latest.max_yr
            ) fr ON ins.company_id = fr.company_id
            LEFT JOIN (
                SELECT mc_inner.*
                FROM market_cap mc_inner
                JOIN (
                    SELECT company_id, MAX(year) as max_yr FROM market_cap GROUP BY company_id
                ) latest_mc ON mc_inner.company_id = latest_mc.company_id AND mc_inner.year = latest_mc.max_yr
            ) mc ON ins.company_id = mc.company_id
            ORDER BY ins.investment_score DESC
        """
        inv_df = pd.read_sql_query(inv_sql, conn)

        # 2. Financial Health Query
        fh_sql = """
            SELECT c.id as Ticker, c.company_name as Company, fh.year as Year,
                   fh.altman_z_score, fh.financial_health_rating, fh.beneish_m_score, fh.manipulation_risk_flag
            FROM financial_health fh
            JOIN companies c ON fh.company_id = c.id
            ORDER BY c.id, fh.year DESC
        """
        fh_df = pd.read_sql_query(fh_sql, conn)

        # 3. Valuation Analysis Query
        val_sql = """
            SELECT c.id as Ticker, c.company_name as Company, vm.year as Year,
                   vm.valuation_score, vm.earnings_yield, vm.fcf_yield, vm.peg_ratio, vm.ev_sales, vm.ev_ebitda, vm.intrinsic_value_score
            FROM valuation_metrics vm
            JOIN companies c ON vm.company_id = c.id
            ORDER BY vm.valuation_score DESC
        """
        val_df = pd.read_sql_query(val_sql, conn)

        # 4. DuPont Analysis Query
        dp_sql = """
            SELECT c.id as Ticker, c.company_name as Company, dp.year as Year,
                   dp.net_profit_margin, dp.asset_turnover, dp.equity_multiplier, dp.dupont_roe
            FROM dupont_analysis dp
            JOIN companies c ON dp.company_id = c.id
            ORDER BY c.id, dp.year DESC
        """
        dp_df = pd.read_sql_query(dp_sql, conn)

        # 5. Risk Flags Query
        rf_sql = """
            SELECT c.id as Ticker, c.company_name as Company, fh.year as Year,
                   fh.financial_health_rating, fh.altman_z_score, fh.manipulation_risk_flag, fh.beneish_m_score
            FROM financial_health fh
            JOIN companies c ON fh.company_id = c.id
            WHERE fh.financial_health_rating = 'Distress Zone' OR fh.manipulation_risk_flag = 'High Risk'
            ORDER BY c.id, fh.year DESC
        """
        rf_df = pd.read_sql_query(rf_sql, conn)

        conn.close()

        output_path = os.path.join(self.output_dir, output_filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        top20_font = Font(name="Calibri", size=11, bold=True)

        sheets_data = [
            ("Investment Ranking", inv_df),
            ("Financial Health", fh_df),
            ("Valuation Analysis", val_df),
            ("DuPont Analysis", dp_df),
            ("Risk Flags", rf_df)
        ]

        for s_title, df_data in sheets_data:
            ws = wb.create_sheet(title=s_title)
            headers = list(df_data.columns)
            ws.append(headers)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r_idx, row in df_data.iterrows():
                row_vals = [row[c] for c in headers]
                ws.append(row_vals)
                curr_row = ws.max_row

                if s_title == "Investment Ranking":
                    # Highlight top 20 investment opportunities
                    if r_idx < 20:
                        for col_idx in range(1, len(headers) + 1):
                            ws.cell(row=curr_row, column=col_idx).font = top20_font
                    
                    rating = str(row.get("investment_rating", ""))
                    if rating in ["Strong Buy", "Buy"]:
                        ws.cell(row=curr_row, column=5).fill = green_fill
                    elif rating == "Hold":
                        ws.cell(row=curr_row, column=5).fill = yellow_fill
                    else:
                        ws.cell(row=curr_row, column=5).fill = red_fill

                elif s_title == "Financial Health" or s_title == "Risk Flags":
                    rating = str(row.get("financial_health_rating", ""))
                    m_flag = str(row.get("manipulation_risk_flag", ""))

                    if rating == "Safe Zone":
                        ws.cell(row=curr_row, column=5 if s_title == "Financial Health" else 4).fill = green_fill
                    elif rating == "Grey Zone":
                        ws.cell(row=curr_row, column=5 if s_title == "Financial Health" else 4).fill = yellow_fill
                    elif rating == "Distress Zone":
                        ws.cell(row=curr_row, column=5 if s_title == "Financial Health" else 4).fill = red_fill

                    if m_flag == "High Risk":
                        ws.cell(row=curr_row, column=7 if s_title == "Financial Health" else 6).fill = red_fill
                    elif m_flag == "Low Risk":
                        ws.cell(row=curr_row, column=7 if s_title == "Financial Health" else 6).fill = green_fill

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(output_path)
        logger.info(f"Exported Investment Intelligence Excel Report to {output_path}.")
        return output_path


if __name__ == "__main__":
    engine = InvestmentScoreEngine()
    engine.run()
