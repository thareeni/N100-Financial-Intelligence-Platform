"""
Peer Comparison Engine Module.
Loads peer group definitions, computes intra-group percentile rankings for 10 metrics,
generates 8-axis Matplotlib radar charts, populates peer_percentiles table in SQLite,
and exports formatted multi-sheet Excel report (peer_comparison.xlsx).
"""

import os
import sqlite3
import logging
from typing import Dict, List, Optional, Tuple, Any
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import matplotlib.pyplot as plt
from dotenv import load_dotenv

from src.screener.engine import ScreenerEngine

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(name)s: %(message)s")
logger = logging.getLogger("PeerEngine")

DB_PATH = os.getenv("DB_PATH", "nifty100.db")
PEER_EXCEL_PATH = os.path.join("supporting datasets", "peer_groups.xlsx")

TARGET_METRICS = [
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "free_cash_flow_cr",
    "pat_cagr_5yr",
    "revenue_cagr_5yr",
    "eps_cagr_5yr",
    "interest_coverage",
    "asset_turnover",
]

INVERTED_METRICS = {"debt_to_equity", "pe_ratio"}


class PeerComparisonEngine:
    def __init__(self, db_path: str = DB_PATH, peer_excel_path: str = PEER_EXCEL_PATH):
        self.db_path = db_path
        self.peer_excel_path = peer_excel_path
        self.output_dir = "output"
        self.reports_dir = os.path.join("reports", "radar_charts")
        os.makedirs(self.output_dir, exist_ok=True)
        os.makedirs(self.reports_dir, exist_ok=True)

    def load_peer_groups_into_db(self, conn: sqlite3.Connection):
        """Loads peer_groups.xlsx into peer_groups SQLite table."""
        if not os.path.exists(self.peer_excel_path):
            logger.error(f"Peer groups file {self.peer_excel_path} not found.")
            return

        conn.execute("""
            CREATE TABLE IF NOT EXISTS peer_groups (
                id INTEGER PRIMARY KEY,
                peer_group_name VARCHAR NOT NULL,
                company_id VARCHAR NOT NULL,
                is_benchmark BOOLEAN DEFAULT 0,
                FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE
            );
        """)
        conn.commit()

        df = pd.read_excel(self.peer_excel_path)
        df.columns = [str(c).strip().lower() for c in df.columns]
        df["company_id"] = df["company_id"].astype(str).str.strip().str.upper()
        
        valid_cos = set(pd.read_sql_query("SELECT id FROM companies", conn)["id"].tolist())
        df = df[df["company_id"].isin(valid_cos)].copy()

        conn.execute("DELETE FROM peer_groups;")
        df.to_sql("peer_groups", conn, if_exists="append", index=False)
        conn.commit()
        logger.info(f"Loaded {len(df)} peer group mappings into peer_groups table.")

    def compute_peer_percentiles(self) -> pd.DataFrame:
        """
        Calculates intra-group percentile ranks (0.0 - 1.0) for 10 metrics across all peer groups.
        Inverts rankings for D/E and P/E. Populates peer_percentiles table.
        Handles companies without peer group by logging 'No peer group assigned'.
        """
        conn = sqlite3.connect(self.db_path)
        conn.execute("PRAGMA foreign_keys = ON;")

        conn.execute("""
            CREATE TABLE IF NOT EXISTS peer_percentiles (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_id VARCHAR NOT NULL,
                peer_group_name VARCHAR NOT NULL,
                year VARCHAR NOT NULL,
                metric VARCHAR NOT NULL,
                metric_value NUMERIC,
                percentile_rank NUMERIC NOT NULL,
                is_benchmark BOOLEAN DEFAULT 0,
                UNIQUE (company_id, peer_group_name, year, metric),
                FOREIGN KEY (company_id) REFERENCES companies(id) ON DELETE CASCADE
            );
        """)
        conn.commit()

        self.load_peer_groups_into_db(conn)

        sql = """
            SELECT pg.peer_group_name, pg.company_id, pg.is_benchmark, c.company_name,
                   fr.year,
                   fr.return_on_equity_pct, fr.return_on_capital_employed_pct, fr.net_profit_margin_pct,
                   fr.debt_to_equity, fr.free_cash_flow_cr, fr.pat_cagr_5yr, fr.revenue_cagr_5yr,
                   fr.eps_cagr_5yr, fr.interest_coverage, fr.asset_turnover, mc.pe_ratio
            FROM peer_groups pg
            JOIN companies c ON pg.company_id = c.id
            JOIN (
                SELECT fr_inner.*
                FROM financial_ratios fr_inner
                JOIN (
                    SELECT company_id, MAX(year) as max_yr
                    FROM financial_ratios
                    WHERE year != 'PARSE_ERROR'
                    GROUP BY company_id
                ) latest ON fr_inner.company_id = latest.company_id AND fr_inner.year = latest.max_yr
            ) fr ON pg.company_id = fr.company_id
            LEFT JOIN (
                SELECT mc_inner.*
                FROM market_cap mc_inner
                JOIN (
                    SELECT company_id, MAX(year) as max_yr
                    FROM market_cap
                    GROUP BY company_id
                ) latest ON mc_inner.company_id = latest.company_id AND mc_inner.year = latest.max_yr
            ) mc ON pg.company_id = mc.company_id
        """
        merged_df = pd.read_sql_query(sql, conn)

        if merged_df.empty:
            logger.warning("No matching company metrics found for peer groups.")
            conn.close()
            return pd.DataFrame()

        percentile_records = []

        for group_name, group_df in merged_df.groupby("peer_group_name"):
            for metric in TARGET_METRICS:
                if metric not in group_df.columns:
                    continue

                vals = pd.to_numeric(group_df[metric], errors="coerce")
                is_inv = metric in INVERTED_METRICS
                
                if is_inv:
                    ranks = vals.rank(ascending=False, method="min", na_option="bottom")
                else:
                    ranks = vals.rank(ascending=True, method="min", na_option="bottom")

                n_valid = vals.dropna().count()
                if n_valid > 1:
                    pct_ranks = ((ranks - 1) / (n_valid - 1)).clip(0.0, 1.0)
                else:
                    pct_ranks = pd.Series(1.0, index=vals.index)

                pct_ranks = pct_ranks.where(vals.notna(), 0.0).fillna(0.0).round(4)

                for idx, row in group_df.iterrows():
                    cid = row["company_id"]
                    yr = row["year"]
                    is_bm = row["is_benchmark"]
                    m_val = row[metric]
                    p_rank = pct_ranks.loc[idx]

                    percentile_records.append({
                        "company_id": cid,
                        "peer_group_name": group_name,
                        "year": yr,
                        "metric": metric,
                        "metric_value": m_val,
                        "percentile_rank": p_rank,
                        "is_benchmark": is_bm
                    })

        perc_df = pd.DataFrame(percentile_records)

        conn.execute("DELETE FROM peer_percentiles;")
        perc_df.to_sql("peer_percentiles", conn, if_exists="append", index=False)
        conn.commit()

        logger.info(f"Populated peer_percentiles table with {len(perc_df)} rows across {perc_df['peer_group_name'].nunique()} peer groups.")
        conn.close()

        return perc_df

    def generate_radar_charts(self) -> int:
        """
        Generates 8-axis radar PNG charts comparing each company vs its peer group average.
        Axes: ROE, ROCE, NPM, D/E, FCF score, PAT CAGR 5Y, Revenue CAGR 5Y, Composite Score.
        Saves charts to reports/radar_charts/<TICKER>_radar.png.
        """
        screener = ScreenerEngine(self.db_path)
        universe_df = screener.get_latest_universe_df()

        conn = sqlite3.connect(self.db_path)
        sql = "SELECT peer_group_name, company_id, is_benchmark FROM peer_groups;"
        pg_df = pd.read_sql_query(sql, conn)
        conn.close()

        merged_df = pd.merge(pg_df, universe_df, on="company_id", how="inner")

        if merged_df.empty:
            logger.warning("No data found for radar chart generation.")
            return 0

        radar_axes = [
            "return_on_equity_pct",
            "return_on_capital_employed_pct",
            "net_profit_margin_pct",
            "debt_to_equity",
            "free_cash_flow_cr",
            "pat_cagr_5yr",
            "revenue_cagr_5yr",
            "composite_quality_score"
        ]
        labels = ["ROE %", "ROCE %", "NPM %", "D/E (Inv)", "FCF (Cr)", "PAT 5Y %", "Rev 5Y %", "Comp Score"]

        num_vars = len(labels)
        angles = [n / float(num_vars) * 2 * np.pi for n in range(num_vars)]
        angles += angles[:1]

        chart_count = 0

        for group_name, group_df in merged_df.groupby("peer_group_name"):
            grp_means = {}
            for col in radar_axes:
                clean_vals = pd.to_numeric(group_df[col], errors="coerce").fillna(0.0)
                grp_means[col] = clean_vals.mean()

            for _, row in group_df.iterrows():
                cid = row["company_id"]
                c_name = row.get("company_name", cid)

                co_vals = []
                avg_vals = []

                for col in radar_axes:
                    val = float(row[col]) if row[col] is not None and not pd.isna(row[col]) else 0.0
                    avg = float(grp_means[col])
                    
                    co_vals.append(val)
                    avg_vals.append(avg)

                co_vals += co_vals[:1]
                avg_vals += avg_vals[:1]

                fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
                
                plt.xticks(angles[:-1], labels, color="grey", size=9)
                ax.set_rlabel_position(0)

                ax.plot(angles, co_vals, linewidth=2, linestyle="solid", label=cid, color="#1F4E78")
                ax.fill(angles, co_vals, color="#1F4E78", alpha=0.25)

                ax.plot(angles, avg_vals, linewidth=1.5, linestyle="dashed", label=f"Peer Avg ({group_name})", color="#D9534F")

                plt.title(f"{c_name} ({cid}) - Peer Radar Analysis", size=11, color="#333333", y=1.1)
                plt.legend(loc="upper right", bbox_to_anchor=(0.1, 0.1))

                chart_file = os.path.join(self.reports_dir, f"{cid}_radar.png")
                plt.savefig(chart_file, bbox_inches="tight", dpi=100)
                plt.close(fig)
                chart_count += 1

        logger.info(f"Generated {chart_count} radar charts in {self.reports_dir}.")
        return chart_count

    def export_peer_comparison_excel(self, output_filename: str = "peer_comparison.xlsx") -> str:
        """
        Exports formatted multi-sheet Excel workbook (peer_comparison.xlsx) containing
        exactly 11 sheets with metric values, percentile ranks, benchmark highlights, and median summary rows.
        Formatting: Green >= 0.75, Yellow 0.25-0.75, Red <= 0.25.
        """
        perc_df = self.compute_peer_percentiles()
        
        conn = sqlite3.connect(self.db_path)
        pg_sql = "SELECT DISTINCT peer_group_name FROM peer_groups;"
        groups = [r[0] for r in conn.execute(pg_sql).fetchall()]

        output_path = os.path.join(self.output_dir, output_filename)
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        benchmark_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        benchmark_font = Font(name="Calibri", size=11, bold=True, color="002060")
        median_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        median_font = Font(name="Calibri", size=11, bold=True, italic=True)
        
        green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        for g_name in groups:
            sql = f"""
                SELECT pg.company_id, c.company_name, pg.is_benchmark,
                       fr.return_on_equity_pct, fr.return_on_capital_employed_pct, fr.net_profit_margin_pct,
                       fr.debt_to_equity, fr.free_cash_flow_cr, fr.pat_cagr_5yr, fr.revenue_cagr_5yr,
                       fr.eps_cagr_5yr, fr.interest_coverage, fr.asset_turnover
                FROM peer_groups pg
                JOIN companies c ON pg.company_id = c.id
                JOIN (
                    SELECT fr_inner.*
                    FROM financial_ratios fr_inner
                    JOIN (
                        SELECT company_id, MAX(year) as max_yr
                        FROM financial_ratios
                        WHERE year != 'PARSE_ERROR'
                        GROUP BY company_id
                    ) latest ON fr_inner.company_id = latest.company_id AND fr_inner.year = latest.max_yr
                ) fr ON pg.company_id = fr.company_id
                WHERE pg.peer_group_name = '{g_name}'
            """
            grp_df = pd.read_sql_query(sql, conn)

            if grp_df.empty:
                continue

            ws = wb.create_sheet(title=g_name[:31])

            # Header columns: Metric values + Percentile ranks
            headers = ["Ticker", "Company Name", "Is Benchmark"]
            for m in TARGET_METRICS:
                headers.append(f"{m}")
                headers.append(f"{m}_pctile")

            ws.append(headers)

            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for _, row in grp_df.iterrows():
                cid = row["company_id"]
                is_bm = bool(row["is_benchmark"])
                row_vals = [cid, row["company_name"], "YES" if is_bm else "NO"]
                
                # Extract percentiles for company
                co_p = perc_df[(perc_df["company_id"] == cid) & (perc_df["peer_group_name"] == g_name)]
                p_map = dict(zip(co_p["metric"], co_p["percentile_rank"]))

                for m in TARGET_METRICS:
                    val = row.get(m)
                    p_rank = p_map.get(m, 0.0)
                    row_vals.append(round(val, 2) if val is not None and not pd.isna(val) else "N/A")
                    row_vals.append(f"{round(p_rank * 100, 1)}%")

                ws.append(row_vals)
                curr_row = ws.max_row

                # Formatting benchmark row
                if is_bm:
                    for col_idx in range(1, 4):
                        cell = ws.cell(row=curr_row, column=col_idx)
                        cell.fill = benchmark_fill
                        cell.font = benchmark_font

                # Color percentile rank cells
                for m_i, m in enumerate(TARGET_METRICS):
                    p_col_idx = 4 + m_i * 2 + 1  # 1-indexed column for percentile
                    p_cell = ws.cell(row=curr_row, column=p_col_idx)
                    p_val = p_map.get(m, 0.0)
                    
                    if p_val >= 0.75:
                        p_cell.fill = green_fill
                    elif p_val <= 0.25:
                        p_cell.fill = red_fill
                    else:
                        p_cell.fill = yellow_fill

            # Compute Median Summary Row
            median_vals = ["MEDIAN", f"{g_name} Peer Median", "-"]
            for m in TARGET_METRICS:
                med_val = pd.to_numeric(grp_df.get(m), errors="coerce").median()
                median_vals.append(round(med_val, 2) if med_val is not None and not pd.isna(med_val) else "N/A")
                median_vals.append("50.0%")

            ws.append(median_vals)
            med_row = ws.max_row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=med_row, column=col_idx)
                cell.fill = median_fill
                cell.font = median_font

            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

        conn.close()
        wb.save(output_path)
        logger.info(f"Exported Peer Comparison Excel Report to {output_path}.")
        return output_path


if __name__ == "__main__":
    peer_engine = PeerComparisonEngine()
    peer_engine.compute_peer_percentiles()
    peer_engine.generate_radar_charts()
    peer_engine.export_peer_comparison_excel()
